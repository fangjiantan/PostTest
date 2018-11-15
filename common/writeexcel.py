# coding:utf-8

from openpyxl import load_workbook
import openpyxl

def copy_excel(excelpath1, excelpath2):
    wb2 = openpyxl.Workbook()
    wb2.save(excelpath2)

    wb1 = openpyxl.load_workbook(excelpath1)
    wb2 = openpyxl.load_workbook(excelpath2)
    sheets1 = wb1.sheetnames
    sheets2 = wb2.sheetnames
    sheet1 = wb1[sheets1[0]]
    sheet2 = wb2[sheets2[0]]
    max_row = sheet1.max_row
    max_column = sheet1.max_column
    for m in list(range(1,max_row+1)):
        for n in list(range(97,97+max_column)):
            n = chr(n)
            i = '%s%d'%(n,m)
            cell1 = sheet1[i].value
            sheet2[i].value = cell1
    wb2.save(excelpath2)
    wb1.close()
    wb2.close()

class Write_excel(object):
    def __init__(self,filename):
        self.filename = filename
        self.wb = load_workbook(self.filename)
        self.ws = self.wb.active
    def write(self,row_n,col_n,value):
        self.ws.cell(row_n,col_n).value = value
        self.wb.save(self.filename)

if __name__ == "__main__":
    copy_excel("C:\\Users\\Administrator\\PycharmProjects\\PostTest\\Testcase\\demo_api.xlsx","C:\\Users\\Administrator\\PycharmProjects\\PostTest\\report\\testreport.xlsx")
    wt = Write_excel("C:\\Users\\Administrator\\PycharmProjects\\PostTest\\report\\testreport.xlsx")
    wt.write(3,1,"Helleop")



